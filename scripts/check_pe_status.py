"""
Check PE status of stocks from Master_Stocks.xlsx against the Trade Automation DB.
Connects via SSH to the automation server, queries PostgreSQL via docker exec.

Uses multi-strategy symbol matching:
  1. Direct match (exact symbol)
  2. No-space match (remove spaces/& from Excel symbol)
  3. Stocks table bridge (symbol, nse_symbol, bse_token)
  4. BSE code reverse lookup (quarterly_results stored as BSE code → resolve via stocks table)
  5. First-word prefix match (first word of Excel symbol matches start of DB symbol)
  6. Company name partial match (quarterly_results.company_name contains Excel symbol words)

Reports: PE Reviewed, PE Pending, In Stocks (no QR), Not Found.
"""

import sys
from pathlib import Path

import paramiko
import openpyxl
from openpyxl.styles import Font, PatternFill


SSH_HOST = "122.165.113.41"
SSH_USER = "kishore"
SSH_PASS = "root"
DB_CONTAINER = "trade_postgres"
DB_NAME = "automation_trade"
DB_USER = "trade_user"


def ssh_exec(client: paramiko.SSHClient, cmd: str) -> str:
    _, stdout, stderr = client.exec_command(cmd, timeout=120)
    out = stdout.read().decode("utf-8", errors="replace")
    err = stderr.read().decode("utf-8", errors="replace")
    if stderr.channel.recv_exit_status() != 0 and not out:
        raise RuntimeError(f"Command failed: {err}")
    return out


def run_db_query(client: paramiko.SSHClient, sql: str) -> str:
    escaped_sql = sql.replace("'", "'\\''")
    cmd = f"docker exec {DB_CONTAINER} psql -U {DB_USER} -d {DB_NAME} -t -A -F '|' -c '{escaped_sql}'"
    return ssh_exec(client, cmd)


def load_excel_stocks(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Master Stocks"]
    stocks = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        sno = row[0].value
        sector = row[1].value
        symbol = row[2].value
        if symbol:
            stocks.append({
                "sno": sno,
                "sector": str(sector).strip() if sector else "",
                "symbol": str(symbol).strip().upper(),
            })
    wb.close()
    return stocks


def build_db_context(client: paramiko.SSHClient) -> dict:
    """
    Fetches all necessary data from the server DB in bulk and builds
    comprehensive lookup structures for symbol resolution.
    """
    # 1. Stocks table: symbol, nse_symbol, bse_token
    raw = run_db_query(client, "SELECT UPPER(symbol), UPPER(COALESCE(nse_symbol,'')), COALESCE(bse_token,0) FROM stocks WHERE symbol IS NOT NULL")
    sym_set = set()
    nospace_to_syms = {}  # nospace -> set of symbols
    bse_to_sym = {}       # bse_token -> symbol
    sym_to_bse = {}       # symbol -> bse_token

    for line in raw.strip().split("\n"):
        if not line.strip():
            continue
        parts = line.strip().split("|")
        if len(parts) >= 3:
            sym = parts[0].strip()
            nse = parts[1].strip()
            try:
                bse = int(parts[2].strip())
            except ValueError:
                bse = 0
            sym_set.add(sym)
            # Index by no-space version of symbol
            ns = sym.replace(" ", "").replace("&", "")
            nospace_to_syms.setdefault(ns, set()).add(sym)
            # Also index NSE symbol
            if nse:
                nospace_to_syms.setdefault(nse.replace(" ", ""), set()).add(sym)
            if bse:
                bse_to_sym[bse] = sym
                sym_to_bse[sym] = bse

    # 2. Quarterly results: stock_symbol + status (reviewed or pending)
    reviewed_raw = run_db_query(client, """
        SELECT DISTINCT UPPER(stock_symbol)
        FROM quarterly_results
        WHERE valuation IS NOT NULL AND valuation != ''
          AND (extraction_status = 'completed' OR user_reviewed = TRUE)
    """)
    reviewed_qr = set(s.strip() for s in reviewed_raw.strip().split("\n") if s.strip())

    pending_raw = run_db_query(client, """
        SELECT DISTINCT UPPER(stock_symbol)
        FROM quarterly_results
        WHERE (valuation IS NULL OR valuation = '')
    """)
    pending_qr = set(s.strip() for s in pending_raw.strip().split("\n") if s.strip())

    # 3. Company names from quarterly_results for fuzzy matching
    company_raw = run_db_query(client, """
        SELECT DISTINCT UPPER(stock_symbol), UPPER(COALESCE(company_name, ''))
        FROM quarterly_results
    """)
    sym_to_company = {}
    company_words_to_syms = {}  # word -> set of stock_symbols
    for line in company_raw.strip().split("\n"):
        if not line.strip():
            continue
        parts = line.strip().split("|")
        if len(parts) >= 2:
            qr_sym = parts[0].strip()
            cname = parts[1].strip()
            sym_to_company[qr_sym] = cname
            for word in cname.split():
                word_clean = word.strip(".-,()&$/")
                if len(word_clean) >= 3:
                    company_words_to_syms.setdefault(word_clean, set()).add(qr_sym)

    # Expand reviewed/pending by resolving BSE codes
    all_qr_syms = reviewed_qr | pending_qr
    resolved_reviewed = set(reviewed_qr)
    resolved_pending = set(pending_qr)

    for qr_sym in all_qr_syms:
        try:
            bse_code = int(qr_sym)
            if bse_code in bse_to_sym:
                actual_sym = bse_to_sym[bse_code]
                if qr_sym in reviewed_qr:
                    resolved_reviewed.add(actual_sym)
                if qr_sym in pending_qr:
                    resolved_pending.add(actual_sym)
        except ValueError:
            pass

    # Also add no-space versions of all resolved symbols
    for s in list(resolved_reviewed):
        resolved_reviewed.add(s.replace(" ", "").replace("&", ""))
    for s in list(resolved_pending):
        resolved_pending.add(s.replace(" ", "").replace("&", ""))

    return {
        "sym_set": sym_set,
        "nospace_to_syms": nospace_to_syms,
        "bse_to_sym": bse_to_sym,
        "sym_to_bse": sym_to_bse,
        "resolved_reviewed": resolved_reviewed,
        "resolved_pending": resolved_pending,
        "reviewed_qr_raw": reviewed_qr,
        "pending_qr_raw": pending_qr,
        "company_words_to_syms": company_words_to_syms,
    }


def check_qr_status(candidates: set, ctx: dict) -> str | None:
    """Check if any candidate symbol is in reviewed or pending QR sets."""
    for c in candidates:
        if c in ctx["resolved_reviewed"]:
            return "REVIEWED"
    for c in candidates:
        if c in ctx["resolved_pending"]:
            return "PENDING"
    return None


def match_symbol(excel_sym: str, ctx: dict) -> str:
    """
    Multi-strategy symbol matching. Returns status string.
    """
    s = excel_sym.strip().upper()
    nospace = s.replace(" ", "").replace("&", "")
    words = s.split()
    first_word = words[0] if words else s

    # Strategy 1: Direct / no-space match against QR
    for candidate in (s, nospace):
        if candidate in ctx["resolved_reviewed"]:
            return "REVIEWED"
        if candidate in ctx["resolved_pending"]:
            return "PENDING"

    # Strategy 2: Stocks table nospace lookup → then check QR
    if nospace in ctx["nospace_to_syms"]:
        matched_syms = ctx["nospace_to_syms"][nospace]
        candidates = set(matched_syms)
        for ms in matched_syms:
            if ms in ctx["sym_to_bse"]:
                candidates.add(str(ctx["sym_to_bse"][ms]))
        status = check_qr_status(candidates, ctx)
        if status:
            return status
        return "IN_STOCKS"

    # Strategy 3: First-word prefix match against stocks table
    # e.g. "TINNA RUBBER" → first word "TINNA" → matches "TINNARUBR"
    prefix_matches = set()
    for db_sym in ctx["sym_set"]:
        db_nospace = db_sym.replace(" ", "")
        if len(first_word) >= 3 and db_nospace.startswith(first_word):
            prefix_matches.add(db_sym)
        elif len(nospace) >= 5 and db_nospace.startswith(nospace[:5]):
            prefix_matches.add(db_sym)

    if prefix_matches:
        candidates = set(prefix_matches)
        for ms in prefix_matches:
            candidates.add(ms.replace(" ", ""))
            if ms in ctx["sym_to_bse"]:
                candidates.add(str(ctx["sym_to_bse"][ms]))
        status = check_qr_status(candidates, ctx)
        if status:
            return status
        return "IN_STOCKS"

    # Strategy 4: Company name word matching
    # e.g. "JASCH" → company_name "JASCH INDUSTRIES..." → matches
    # Use all words from Excel symbol to find QR entries
    matched_qr_syms = None
    for word in words:
        word_clean = word.strip(".-,()&$/")
        if len(word_clean) < 3:
            continue
        word_matches = ctx["company_words_to_syms"].get(word_clean, set())
        if matched_qr_syms is None:
            matched_qr_syms = set(word_matches)
        else:
            # Intersect: all words must match
            intersection = matched_qr_syms & word_matches
            if intersection:
                matched_qr_syms = intersection

    if matched_qr_syms:
        candidates = set(matched_qr_syms)
        for ms in list(matched_qr_syms):
            candidates.add(ms.replace(" ", ""))
        status = check_qr_status(candidates, ctx)
        if status:
            return status

    # Strategy 5: Check if exact symbol or nospace is in stocks table (no QR data)
    if s in ctx["sym_set"] or nospace in ctx["sym_set"]:
        return "IN_STOCKS"

    # Strategy 6: Reverse BSE lookup — check if any BSE code for prefix-matched stocks is in QR
    for db_sym in ctx["sym_set"]:
        if db_sym.startswith(first_word) and len(first_word) >= 4:
            if db_sym in ctx["sym_to_bse"]:
                bse_str = str(ctx["sym_to_bse"][db_sym])
                if bse_str in ctx["resolved_reviewed"]:
                    return "REVIEWED"
                if bse_str in ctx["resolved_pending"]:
                    return "PENDING"

    return "NOT FOUND"


def save_excel_with_status(filepath: str, stocks: list[dict], status_map: dict):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Master Stocks"]

    status_col = 5
    ws.cell(row=1, column=status_col, value="PE Status")

    fills = {
        "REVIEWED": (PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                     Font(color="006100", bold=True)),
        "PENDING": (PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                    Font(color="9C5700", bold=True)),
        "IN_STOCKS": (PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
                      Font(color="1F4E79", bold=True)),
        "NOT FOUND": (PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                      Font(color="9C0006", bold=True)),
    }

    for i, stock in enumerate(stocks, start=2):
        status = status_map.get(stock["symbol"], "NOT FOUND")
        cell = ws.cell(row=i, column=status_col)
        cell.value = status
        fill, font = fills.get(status, fills["NOT FOUND"])
        cell.fill = fill
        cell.font = font

    ws.column_dimensions["E"].width = 14

    output_path = filepath.replace(".xlsx", "_PE_Status.xlsx")
    try:
        wb.save(output_path)
    except PermissionError:
        alt_path = filepath.replace(".xlsx", "_PE_Status_new.xlsx")
        print(f"  [WARN] {output_path} is locked, saving as {alt_path}")
        wb.save(alt_path)
        output_path = alt_path
    wb.close()
    return output_path


def main():
    excel_path = Path(__file__).parent.parent / "Master_Stocks (8).xlsx"
    if not excel_path.exists():
        print(f"ERROR: {excel_path} not found")
        sys.exit(1)

    print(f"Loading stocks from: {excel_path}")
    stocks = load_excel_stocks(str(excel_path))
    print(f"Total stocks in Excel: {len(stocks)}")

    print(f"\nConnecting to {SSH_HOST}...")
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(SSH_HOST, username=SSH_USER, password=SSH_PASS, timeout=15)
    print("Connected!")

    print("Building DB context (stocks table + quarterly_results + company names)...")
    ctx = build_db_context(client)
    print(f"  Stocks table: {len(ctx['sym_set'])} symbols, {len(ctx['bse_to_sym'])} BSE mappings")
    print(f"  QR Reviewed (raw): {len(ctx['reviewed_qr_raw'])}, Pending (raw): {len(ctx['pending_qr_raw'])}")
    print(f"  QR Reviewed (resolved): {len(ctx['resolved_reviewed'])}, Pending (resolved): {len(ctx['resolved_pending'])}")
    print(f"  Company name index: {len(ctx['company_words_to_syms'])} unique words")

    client.close()

    # Classify each stock
    status_map = {}
    counts = {"REVIEWED": 0, "PENDING": 0, "IN_STOCKS": 0, "NOT FOUND": 0}
    sector_stats = {}
    classified = {"REVIEWED": [], "PENDING": [], "IN_STOCKS": [], "NOT FOUND": []}

    for stock in stocks:
        status = match_symbol(stock["symbol"], ctx)
        status_map[stock["symbol"]] = status
        counts[status] += 1
        classified[status].append(stock)
        sec = stock["sector"]
        if sec not in sector_stats:
            sector_stats[sec] = {"REVIEWED": 0, "PENDING": 0, "IN_STOCKS": 0, "NOT FOUND": 0, "total": 0}
        sector_stats[sec][status] += 1
        sector_stats[sec]["total"] += 1

    # Print Report
    total = len(stocks)
    print(f"\n{'=' * 80}")
    print("PE STATUS REPORT")
    print(f"{'=' * 80}")
    print(f"\n  {'Category':<40} {'Count':>6} {'Pct':>8}")
    print(f"  {'-' * 56}")
    print(f"  {'[REVIEWED] PE Done':<40} {counts['REVIEWED']:>6} {counts['REVIEWED']/total*100:>7.1f}%")
    print(f"  {'[PENDING] PE Pending':<40} {counts['PENDING']:>6} {counts['PENDING']/total*100:>7.1f}%")
    print(f"  {'[IN_STOCKS] In stocks, no QR data':<40} {counts['IN_STOCKS']:>6} {counts['IN_STOCKS']/total*100:>7.1f}%")
    print(f"  {'[NOT FOUND] Not in any table':<40} {counts['NOT FOUND']:>6} {counts['NOT FOUND']/total*100:>7.1f}%")
    print(f"  {'-' * 56}")
    print(f"  {'Total':<40} {total:>6}")

    # Sector breakdown
    print(f"\n{'=' * 80}")
    print("SECTOR BREAKDOWN")
    print(f"{'=' * 80}")
    print(f"  {'Sector':<22} {'Total':>5} {'Rvwd':>5} {'Pend':>5} {'InSt':>5} {'Miss':>5} {'Rvwd%':>6}")
    print(f"  {'-' * 55}")
    for sec in sorted(sector_stats, key=lambda s: sector_stats[s]["total"], reverse=True):
        ss = sector_stats[sec]
        pct = ss["REVIEWED"] / ss["total"] * 100 if ss["total"] else 0
        print(f"  {sec:<22} {ss['total']:>5} {ss['REVIEWED']:>5} {ss['PENDING']:>5} {ss['IN_STOCKS']:>5} {ss['NOT FOUND']:>5} {pct:>5.0f}%")

    # List NOT FOUND stocks
    if classified["NOT FOUND"]:
        print(f"\n{'=' * 80}")
        print(f"NOT FOUND STOCKS ({counts['NOT FOUND']})")
        print(f"{'=' * 80}")
        print(f"  {'#':<5} {'Symbol':<28} {'Sector':<20}")
        print(f"  {'-' * 55}")
        for i, s in enumerate(classified["NOT FOUND"], 1):
            print(f"  {i:<5} {s['symbol']:<28} {s['sector']:<20}")

    # List IN_STOCKS
    if classified["IN_STOCKS"]:
        print(f"\n{'=' * 80}")
        print(f"IN STOCKS TABLE BUT NO QUARTERLY RESULTS ({counts['IN_STOCKS']})")
        print(f"{'=' * 80}")
        print(f"  {'#':<5} {'Symbol':<28} {'Sector':<20}")
        print(f"  {'-' * 55}")
        for i, s in enumerate(classified["IN_STOCKS"], 1):
            print(f"  {i:<5} {s['symbol']:<28} {s['sector']:<20}")

    # Save Excel
    output_file = save_excel_with_status(str(excel_path), stocks, status_map)
    print(f"\n{'=' * 80}")
    print(f"Output Excel saved: {output_file}")
    print(f"{'=' * 80}\n")


if __name__ == "__main__":
    main()
